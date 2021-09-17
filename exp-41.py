import openpyxl as xl
from openpyxl.chart import BarChart, Reference
# chart is a module and there are 2  classes BarChart and Reference

wb = xl.load_workbook('transactions.xlSX')
sheet = wb['Sheet1']
cell = sheet["a1"]
cell = sheet.cell(1, 2)
print(cell.value)  # cell is the method here for sheet object

print(sheet.max_row)
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    print(cell.value)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

values = Reference(sheet,
          min_row = 2,
          max_row = sheet.max_row,
          min_col = 4,
          max_col = 4)
# values is instance of reference class
# values store the values in col 4
chart = BarChart()     # chart is an instance of barchart class
chart.add_data(values)
sheet.add_chart(chart, 'e2')


wb.save("transaction2.xlsx")

# +1 because it will consider n-1. In the form of (row, column)
# We dont want to print heading price in the 1st row 3rd column.
# so range starts from 2

